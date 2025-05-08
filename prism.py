import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from odps import ODPS
from datetime import timedelta
import os
from dotenv import load_dotenv

# ====== Streamlit Layout ======

st.markdown("""
<style>
/* Style all input widgets: date_input, selectbox, text_input, etc. */
div[data-baseweb="input"] input,
div[data-baseweb="select"] > div {
    background-color: #001f1f !important;
    color: #00ffae !important;
    border: 1px solid #00ffae !important;
    font-family: 'Courier New', monospace !important;
}

/* Dropdown menu background & hover */
[data-baseweb="select"] .css-1n76uvr-option {
    background-color: #001f1f !important;
    color: #00ffae !important;
}
[data-baseweb="select"] .css-1n76uvr-option:hover {
    background-color: #003333 !important;
}

/* Dropdown selected item styling */
.css-1pahdxg-control,
.css-1s2u09g-control {
    background-color: #001f1f !important;
    color: #00ffae !important;
    border: 1px solid #00ffae !important;
}

/* Adjust label spacing */
.css-1l269bu {
    margin-bottom: 5px;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
    <style>

    
    /* Global dark theme */
    body {
        background-color: #0d1117;
        color: #00ffae;
    }
    .stApp {
        background-color: #0d1117;
        color: #00ffae;
        font-family: 'Courier New', monospace;
    }

    /* Headings */
    h1, h2, h3, h4 {
        color: #00ffae;
        text-shadow: 0 0 5px #00ffae;
    }

    /* Buttons */
    .stButton > button {
        background-color: #001f1f;
        color: #00ffae;
        border: 1px solid #00ffae;
        box-shadow: 0 0 10px #00ffae;
        font-family: 'Courier New', monospace;
    }
    .stButton > button:hover {
        background-color: #003f3f;
        color: #00ffae;
    }

    /* Text input & date picker */
    .stTextInput input, .stDateInput input {
        background-color: #0b1a1a;
        color: #00ffae;
        border: 1px solid #00ffae;
        font-family: 'Courier New', monospace;
    }

    /* Dataframe styling */
    .stDataFrame {
        background-color: #0d1117 !important;
        color: #00ffae !important;
    }

    /* Spinner and alerts */
    .stSpinner, .stAlert {
        color: #00ffae;
        background-color: #001f1f;
        border: 1px solid #00ffae;
        font-family: 'Courier New', monospace;
    }

    /* Download button */
    .stDownloadButton > button {
        background-color: #001f1f;
        color: #00ffae;
        border: 1px solid #00ffae;
        box-shadow: 0 0 10px #00ffae;
        font-family: 'Courier New', monospace;
    }

    /* Markdown sections */
    .markdown-text-container {
        color: #00ffae;
    }
    
    
    </style>
""", unsafe_allow_html=True)


st.markdown("""
    <style>
        h1 {
            margin-bottom: 0px;  /* Reduces the gap below the <h1> */
        }
        h5 {
            margin-top: 0px;  /* Reduces the gap above the <h5> */
        }
    </style>
""", unsafe_allow_html=True)

st.markdown("<h5 style='text-align: center; font-size: 156px; line-height: 0.5; padding-right: 50px;'>👁⃤</h5>", unsafe_allow_html=True)
st.markdown("<h1 style='text-align: center; font-size: 128px; line-height: 0.75;'>P.R.I.S.M</h1>", unsafe_allow_html=True)
st.markdown("<h5 style='text-align: center; font-size: 32px; line-height: 1.0;'>Personalized Reporting Interface for SQL and MaxCompute</h5>", unsafe_allow_html=True)
st.markdown("<hr style='border: 1px solid #00ffae; line-height: 0.75;'>", unsafe_allow_html=True)

st.markdown("<span style='color:#00ffae; font-weight:bold; font-family:Courier New;'>Start Date</span>", unsafe_allow_html=True)
start_date = st.date_input("Select Start Date", label_visibility="collapsed")


st.markdown("<span style='color:#00ffae; font-weight:bold; font-family:Courier New;'>End Date</span>", unsafe_allow_html=True)
end_date = st.date_input("Select End Date", label_visibility="collapsed")


# Custom header styling
st.markdown("<span style='color:#00ffae; font-weight:bold; font-family:Courier New;'>Select Camera ID</span>", unsafe_allow_html=True)

# Custom dropdown (replace the list with your actual camera IDs)
camera_options = ['JA029-77', 'JTAH004-39', 'JSP001-166', 'JC009-17', 'JI005-60', 'JKG036-115', 'JM008-10', 'JPH004-9', 'JP016-166', 'JSAS02-6', 'JT07-42', 'JTR011-209', 'KSH005-13']  # <- Add your list here
camera_id = st.selectbox("Select Camera ID", options=camera_options, label_visibility="collapsed")



if st.button("Run Query"):
    if start_date and end_date and camera_id:
        try:
            load_dotenv()  # Loads from .env

            access_id = os.getenv("MAXCOMPUTE_ACCESS_ID")
            access_key = os.getenv("MAXCOMPUTE_ACCESS_KEY")
            endpoint_URL = os.getenv("ENDPOINT")
            project_name = os.getenv("PROJECT")
            
            # Connect to ODPS
            o = ODPS(access_id, access_key, project=project_name, endpoint=endpoint_URL)
            start_str = start_date.strftime('%Y%m%d')
            end_str = end_date.strftime('%Y%m%d')

            # Load SQL
            with open('MaxComputeQuery.sql', 'r') as file:
                sql_template = file.read()
            sql_query = sql_template.format(start_date=start_str, end_date=end_str, cam_id=camera_id)

            # Query data and store in df
            with st.spinner("Running query and fetching data..."):
                query_job = o.execute_sql(sql_query)
                result = query_job.open_reader(tunnel=True)
                df = result.to_pandas(n_process=1)
                
                df['combined_ai_average_speed'] = ( # clean data : impute missing values in combined_ai_average_speed
                    df
                    .groupby(['lane', 'vehicle_type'])['combined_ai_average_speed']
                    .apply(lambda group: (group.ffill() + group.bfill()) / 2)
                )
                df['density'] = df.apply( # calculate and append density column on queried dataset
                    lambda row: row['combined_avg_volume'] / row['combined_ai_average_speed']
                    if row['combined_ai_average_speed'] > 0 and row['vehicle_type'] == 'ALL' else np.nan,
                    axis=1
                )

            st.markdown("""
                <style>
                
                @keyframes glow {
                    0% { box-shadow: 0 0 5px #00ffae; }
                    50% { box-shadow: 0 0 20px #00ffae; }
                    100% { box-shadow: 0 0 5px #00ffae; }
                }
                .glowing-box {
                    animation: glow 2s infinite;
                }
                </style>
            """, unsafe_allow_html=True)

            st.markdown("""
                <div class="glowing-box" style="
                background-color: #002b2b;
                color: #00ffae;
                padding: 12px;
                border-left: 5px solid #00ffae;
                font-family: 'Courier New', monospace;
                font-weight: bold;
                ">✅ Data fetched successfully! Please wait for Excel report...</div>
                
""", unsafe_allow_html=True)

            st.dataframe(df)

            # Create Excel in memory
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                # =============================================================== Sheet 1: RAW DATA
                df.to_excel(writer, sheet_name='Sheet1', index=False)

                # =============================================================== Sheet 2: DAILY INBOUND/OUTBOUND
                vehicle_order = ['bus', 'car', 'motorbike', 'truck', 'lorry', 'van']
                pivot = df.groupby(['day', 'vehicle_type', 'direction'])['combined_avg_volume'].sum().unstack(fill_value=0).reset_index()
                pivot.rename(columns={'IN': 'Inbound', 'OUT': 'Outbound'}, inplace=True)
                pivot['Total'] = pivot['Inbound'] + pivot['Outbound']
                pivot['vehicle_type'] = pd.Categorical(pivot['vehicle_type'], categories=vehicle_order, ordered=True)
                pivot = pivot.sort_values(by=['day', 'vehicle_type']).reset_index(drop=True)

                formatted_tables = []
                for date, group in pivot.groupby('day'):
                    table = pd.DataFrame({
                        'Veh. Classification': group['vehicle_type'],
                        'Inbound': group['Inbound'],
                        'Outbound': group['Outbound'],
                        'Total': group['Total']
                    }).reset_index(drop=True)
                    header = pd.DataFrame({
                        'Veh. Classification': [date],
                        'Inbound': [''],
                        'Outbound': [''],
                        'Total': ['']
                    })
                    formatted_tables.append(pd.concat([header, table], ignore_index=True))

                final_output = pd.concat(formatted_tables, ignore_index=True)
                final_output.to_excel(writer, sheet_name='Sheet2', index=False)

                # ================================================================ Sheet 3: DAILY VOLUME
                df['day'] = pd.to_datetime(df['day'])
                daily_summary = df.groupby(['day', 'direction'])['combined_avg_volume'].sum().unstack(fill_value=0)
                daily_summary['Total Daily Volume'] = daily_summary.sum(axis=1)
                daily_summary = daily_summary.reset_index()
                daily_summary.rename(columns={'IN': 'Inbound', 'OUT': 'Outbound'}, inplace=True)
                daily_summary.to_excel(writer, sheet_name='Sheet3', index=False)

                # ================================================================ Hourly Sheets per Day
                full_df = df.copy()
                current_date = pd.to_datetime(start_date)
                end_dt = pd.to_datetime(end_date)

                while current_date <= end_dt:
                    day_df = full_df[full_df['day'] == current_date]
                    day_df = day_df[day_df['vehicle_type'] == 'ALL']

                    if not day_df.empty:
                        day_df['datetime'] = pd.to_datetime(day_df['day'].dt.date.astype(str) + ' ' + day_df['hour_s'])
                        day_df['hour'] = day_df['datetime'].dt.hour

                        summary = day_df.groupby(['hour', 'direction'])['combined_avg_volume'].sum().unstack(fill_value=0)
                        for col in ['IN', 'OUT']:
                            if col not in summary.columns:
                                summary[col] = 0

                        final_table = pd.DataFrame({
                            'Time': pd.to_datetime(summary.index, format='%H').strftime('%I:00 %p'),
                            'Volume(IB)': summary['IN'],
                            'Volume(OB)': summary['OUT'],
                        })
                        final_table['Total Volume per Hour'] = final_table['Volume(IB)'] + final_table['Volume(OB)']

                        hours_list = pd.date_range('01:00', '23:00', freq='1H').strftime('%I:00 %p')
                        final_table = final_table.set_index('Time').reindex(hours_list, fill_value=np.nan).reset_index()
                        
                        
                        #===============calculation for LOS(IN) and LOS(OUT) columns here
                        
                        # Build density_avg as before
                        density_avg = day_df.groupby(['hour', 'direction'])['density'].mean().unstack(fill_value=0)

                        # Add missing columns
                        for col in ['IN', 'OUT']:
                            if col not in density_avg.columns:
                                density_avg[col] = 0

                        # Get LOS grades
                        def get_los(density):
                            if density < 6.82:
                                return 'A'
                            elif density < 11.18:
                                return 'B'
                            elif density < 16.15:
                                return 'C'
                            elif density < 21.74:
                                return 'D'
                            elif density < 27.97:
                                return 'E'
                            else:
                                return 'F'

                        # Apply LOS grading
                        los_df = pd.DataFrame({
                            'LOS(IB)': density_avg['IN'].apply(get_los),
                            'LOS(OB)': density_avg['OUT'].apply(get_los)
                        })

                        # Now reindex los_df using actual hours 1–23
                        hours_range = list(range(1, 24))  # 1AM to 11PM
                        los_df = los_df.reindex(hours_range, fill_value=np.nan)

                        # Format Time labels to match final_table
                        los_df.index = pd.to_datetime(los_df.index, format='%H').strftime('%I:00 %p')
                        los_df = los_df.reset_index(drop=True)

                        # Now safely concat
                        final_table = pd.concat([final_table.reset_index(drop=True), los_df], axis=1)
                        
                        # ===================== LOS CALCULATION ENDS ================================
                        
                        sheet_name = f'Hourly_{current_date.strftime("%Y-%m-%d")}'
                        final_table.to_excel(writer, sheet_name=sheet_name[:31], index=False)  # Excel sheet name limit


                    current_date += timedelta(days=1)

            # Reload workbook to insert charts
            workbook = load_workbook(excel_buffer)

            # ===== Chart for Sheet 3
            temp_chart_1 = BytesIO()
            plt.figure(figsize=(10, 5))
            plt.plot(daily_summary['day'], daily_summary['Inbound'], marker='o', label='Inbound')
            plt.plot(daily_summary['day'], daily_summary['Outbound'], marker='o', label='Outbound')
            plt.title('Daily Traffic Volume')
            plt.xlabel('Date')
            plt.ylabel('Volume')
            plt.xticks(daily_summary['day'], rotation=45)
            plt.ylim(bottom=0)
            plt.legend()
            plt.grid(True)
            plt.tight_layout()
            plt.savefig(temp_chart_1, format='png')
            plt.close()
            temp_chart_1.seek(0)
            img_1 = Image(temp_chart_1)
            workbook['Sheet3'].add_image(img_1, 'G2')

            # ===== Charts for Hourly Sheets
            for sheetname in workbook.sheetnames:
                if sheetname.startswith('Hourly_'):
                    sheet = workbook[sheetname]
                    df_sheet = pd.read_excel(excel_buffer, sheet_name=sheetname)
                    
                    temp_chart = BytesIO()
                    plt.figure(figsize=(10, 6))
                    plt.plot(df_sheet['index'], df_sheet['Volume(IB)'], marker='o', label='Inbound Volume', color='blue')
                    plt.plot(df_sheet['index'], df_sheet['Volume(OB)'], marker='o', label='Outbound Volume', color='darkorange')
                    plt.title(f'Hourly Traffic Volume for {sheetname[7:]}')
                    plt.xlabel('Time')
                    plt.ylabel('Volume')
                    plt.xticks(rotation=45)
                    plt.ylim(bottom=0)
                    plt.legend()
                    plt.grid(True)
                    plt.tight_layout()
                    plt.savefig(temp_chart, format='png')
                    plt.close()
                    temp_chart.seek(0)
                    chart_image = Image(temp_chart)
                    sheet.add_image(chart_image, 'G2')

            final_buffer = BytesIO()
            workbook.save(final_buffer)
            final_buffer.seek(0)

            st.download_button(
                label="Download Excel (XLSX)",
                data=final_buffer,
                file_name="combined_output_with_charts.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Failed to fetch data: {e}")
    else:
        st.warning("Please enter valid inputs.")

#Components list :

#st.markdown
#st.subheader
#st.title
#st.text_area
#st.button
#st.warning
#st.spinner
#st.success
#st.dataframe
#st.download_button
#st.error


